import pandas as pd
import sys
import os
import string
import shutil
import time
from dateutil import relativedelta
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Side, Alignment, PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.drawing.image import Image
from openpyxl.formatting import Rule
from openpyxl.utils.dataframe import dataframe_to_rows

las_month = datetime.strftime(datetime.today() - pd.DateOffset(months=1), '%b')
lat_month_results = datetime.strftime(datetime.today() - pd.DateOffset(months=1), '_%y%m_%b')
two_month_results = datetime.strftime(datetime.today() - pd.DateOffset(months=2), '_%y%m_%b')
thr_month_results = datetime.strftime(datetime.today() - pd.DateOffset(months=3), '_%y%m_%b')
for_month_results = datetime.strftime(datetime.today() - pd.DateOffset(months=4), '_%y%m_%b')
fiv_month_results = datetime.strftime(datetime.today() - pd.DateOffset(months=5), '_%y%m_%b')
six_month_results = datetime.strftime(datetime.today() - pd.DateOffset(months=2), '_%y%m_%b')

# Defining folder structure
path_to_report = ['../Reports/{} 2018/A-C/'.format(las_month, ), '../Reports/{} 2018/D-F/'.format(las_month, ),
        '../Reports/{} 2018/G-K/'.format(las_month, ), '../Reports/{} 2018/L-N/'.format(las_month, ),
        '../Reports/{} 2018/O-S/'.format(las_month, ), '../Reports/{} 2018/T-Z/'.format(las_month, )]

# Defining current template
path_to_template = '../Templates/{} Template.xlsx'.format(las_month, )

# Building report structure
for p in path_to_report:
    try:
        os.makedirs(p)
    except OSError:
        if not os.isdir(p):
            raise

# Building report log
sys.stdout = open('../Reports/{} 2018/Report Log.csv'.format(las_month, ), 'w')

# Defining dataset location
filename = '../Data Sets/{} Data Set.xlsx'.format(las_month, )

# Import dataset
overalldf = pd.read_excel(filename, encoding='utf-8').fillna(value='').sort_values(by='Emp_Login')

headercl = list(list(overalldf)[8:20])

# Building a dataframe containing all usernames in the dataset
hierarchyloginlist = overalldf[['Level4_login', 'Level5_login', 'Level6_login',
                       'Level7_login', 'Level8_login',
                       'Level9_login', 'Level10_login',
                       'Level11_login', 'Level12_login', 'Login']]

# building a list of all columns with usernames
masterhierarchyloginlist = list(set(hierarchyloginlist.get_values().flatten().tolist()))

# removing any duplicate usernames
masterusernamelist = list(filter(None, masterhierarchyloginlist))
masterusernamelist.sort()

img = Image('../Templates/Connlogo.png')

result = pd.DataFrame()

# lowercase alphabet list to match with username
alphalist = list(string.ascii_lowercase)

alist = list(string.ascii_uppercase) # uppercase alphabet list to match with Excel headers
alist.extend(('AA', 'AB', 'AC', 'AD', 'AE')) # Extending list to match Excel template

# Slicing alphabet list to separate files into their respective files
longlist = [alphalist[:3], alphalist[3:6], alphalist[6:11], alphalist[11:14], alphalist[14:19], alphalist[19:26]]

for i in masterusernamelist:
    for a in range(0, 6):
        for l in longlist[a]:
            if i.startswith(str(l)):
                # Create a copy of the template for the current leader
                shutil.copy2(path_to_template, path_to_report[a] + '{}.xlsx'.format(i, ))
                result = overalldf[overalldf['Emp_Login'] == ('{}'.format(i, ))]
                for q in range(3, 13):
                    if any(overalldf['Level{}_login'.format(q, )] == ('{}'.format(i, ))):
                        reportdf = overalldf[overalldf['Level{}_login'.format(q, )] == ('{}'.format(i, ))]
                        result = result.append(reportdf)

                # Re-ordering dataframe to match Excel template
                result = result[['Emp_Login', 'Emp_Id', 'Emp_Name', 'Business_Title', 'Job_Level', 'Site_Code',
                                 'MOM_Class_Trend', '{}'.format(lat_month_results, ), '{}'.format(two_month_results, ),
                                 '{}'.format(thr_month_results, ), '{}'.format(for_month_results, ), '{}'.format(fiv_month_results, ),
                                 '{}'.format(six_month_results, ), 'L68_SOC_MAN_', 'L68_SOC_ENG_', 'L68_SOC_SIL_',
                                 'L68_SOC_CUL_', 'L68_DIR_LOM_', 'Country', 'tenure2', 'HC_All', 'HC_Blue', 'HC_Temp',
                                 'Level_4', 'Level_5', 'Level_6', 'Level_7', 'Level_8',
                                 'Level_9', 'Level_10', 'Level_11', 'Level_12']]
                book = load_workbook(path_to_report[a] + '{}.xlsx'.format(i, ))
                writer = pd.ExcelWriter(book, engine='openpyxl')
                ws = book['Report']
                c = ws['G3']
                ws.freeze_panes = c
                ws.row_dimensions[3].hidden = True
                ws1 = book['Intro']
                ws1.add_image(img, 'B2')
                #
                # Writing data to Excel file
                #
                for r in dataframe_to_rows(result, index=False, header=False):
                    ws.cell(column=1, row=2)
                    ws.append(r)
                resultlen = len(result) + 4
                #
                # Adding conditional Formatting
                #
                thin = Side(border_style='thin', color='000000')
                thick = Side(border_style='medium', color='000000')

                for x in range(4, resultlen):
                    for y in alist:
                        sty = ws[y + str(x - 1)]
                        styhierarchytwelve = ws['AF' + str(x - 1)]
                        stylastrow = ws[y + str(resultlen - 1)]
                        stylastcell = ws['AF' + str(resultlen - 1)]
                        sty.alignment = Alignment(horizontal='center', vertical='center')
                        sty.border = Border(left=thin, right=thin, bottom=thin)
                        styhierarchytwelve.alignment = Alignment(horizontal='center', vertical='center')
                        styhierarchytwelve.border = Border(left=thin, right=thick, bottom=thin)
                        stylastrow.alignment = Alignment(horizontal='center', vertical='center')
                        stylastrow.border = Border(left=thin, right=thin, bottom=thick)
                        stylastcell.alignment = Alignment(horizontal='center', vertical='center')
                        stylastcell.border = Border(left=thin, right=thick, bottom=thick)
				#
				# Dashboard heatmap styling
				#
                if any(result[headercl] == 'Field 8'):
                    dxf = DifferentialStyle(font=Font(color='FE0000'))
                    rule = Rule(type='containsText', operator='containsText', text='Field 8', dxf=dxf)
                    rule.formula = ['NOT(ISERROR(SEARCH("Field 8", G1)))']
                    ws.conditional_formatting.add('G:G', rule)
                if any(result[headercl] == 'Field 9'):
                    dxf = DifferentialStyle(font=Font(color='FE0000', bold=True))
                    rule = Rule(type='containsText', operator='containsText', text='Field 9', dxf=dxf)
                    rule.formula = ['NOT(ISERROR(SEARCH("Field 9", G1)))']
                    ws.conditional_formatting.add('G:G', rule)
                if any(result[headercl] == 'Field 10'):
                    dxf = DifferentialStyle(font=Font(color='00FF00'))
                    rule = Rule(type='containsText', operator='containsText', text='Field 10', dxf=dxf)
                    rule.formula = ['NOT(ISERROR(SEARCH("Field 10", G1)))']
                    ws.conditional_formatting.add('G:G', rule)
                if any(result[headercl] == 'Field 11'):
                    dxf = DifferentialStyle(font=Font(color='00FF00', bold=True))
                    rule = Rule(type='containsText', operator='containsText', text='Field 11', dxf=dxf)
                    rule.formula = ['NOT(ISERROR(SEARCH("Field 11", G1)))']
                    ws.conditional_formatting.add('G:G', rule)
                if any(result[headercl] == 'Field 2'):
                    dxf = DifferentialStyle(fill=PatternFill(bgColor='CCCCCC'))
                    rule = Rule(type='containsText', operator='containsText', text='Field 2', dxf=dxf)
                    rule.formula = ['NOT(ISERROR(SEARCH("Field 2", G1)))']
                    ws.conditional_formatting.add('G:R', rule)
                if any(result[headercl] == 'Flat'):
                    dxf = DifferentialStyle(fill=PatternFill(bgColor='FFFFFF'))
                    rule = Rule(type='containsText', operator='containsText', text='Flat', dxf=dxf)
                    rule.formula = ['NOT(ISERROR(SEARCH("Flat", G1)))']
                    ws.conditional_formatting.add('G:G', rule)
                if any(result[headercl] == 'Field 4'):
                    dxf = DifferentialStyle(fill=PatternFill(bgColor='FF9B9B'))
                    rule = Rule(type='containsText', operator='containsText', text='Field 4', dxf=dxf)
                    rule.formula = ['NOT(ISERROR(SEARCH("Field 4", H1)))']
                    ws.conditional_formatting.add('H:R', rule)
                if any(result[headercl] == 'Field 3'):
                    dxf = DifferentialStyle(fill=PatternFill(bgColor='FF0000'))
                    rule = Rule(type='containsText', operator='containsText', text='Field 3', dxf=dxf)
                    rule.formula = ['NOT(ISERROR(SEARCH("Field 3", H1)))']
                    ws.conditional_formatting.add('H:R', rule)
                if any(result[headercl] == 'Field 6'):
                    dxf = DifferentialStyle(fill=PatternFill(bgColor='b3ffb3'))
                    rule = Rule(type='cellIs', operator='equal', text='Field 6', dxf=dxf)
                    rule.formula = ['"Field 6"']
                    ws.conditional_formatting.add('H:R', rule)
                if any(result[headercl] == 'Field 7'):
                    dxf = DifferentialStyle(fill=PatternFill(bgColor='66FF66'))
                    rule = Rule(type='cellIs', operator='equal', text='Field 7', dxf=dxf, priority=1)
                    rule.formula = ['"Field 7"']
                    ws.conditional_formatting.add('H:R', rule)
                #
                # Saving finial workbook
                #
                book.save(path_to_report[a] + '{}.xlsx'.format(i, ))
    print(i + ',' + str(len(result)))
