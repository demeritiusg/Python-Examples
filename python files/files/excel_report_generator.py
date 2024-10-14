"""
to generate a report using mostly pandas
"""

# df = pd.read_csv(r'D:\GIT_REPOs\Python-Examples\Python-Examples\python files\data\african_mobile_data.csv', index_col=0)
# data = df.reset_index(drop=True)
# data.to_csv(r'D:\GIT_REPOs\Python-Examples\Python-Examples\python files\data\african_mobile_data_2.csv', index=False)
# print(df)

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment


data = {
    'name': ['alice', 'bob', 'charlie', 'david', 'eve', 'frank'],
    'department':['hr', 'finance', 'hr', 'it', 'fianace', 'it'],
    'salary':[50000, 60000, 55000, 70000, 62000, 72000]
}

df = pd.DataFrame(data)

group_column = 'department'

unique_values = df[group_column].unique()

for value in unique_values:

    filtered_df = df[df[group_column] == value]

    wb = Workbook()
    ws = wb.active
    ws.title = f"{value} Report"

    #adding data from data from
    for r_idx, row in enumerate(dataframe_to_rows(filtered_df, index=False, header=True), 1):
        for c_idx, cell_velue in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=cell_velue)

            if r_idx == 1:
                cell.fill = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0", fill_type="solid")
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    file_name = f"{value}_report.xlsx"

    wb.save(file_name)

    print(f"report for {value} created as {file_name}")
